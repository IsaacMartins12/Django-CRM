from django.shortcuts import render,redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from .forms import SignUpForm, AddRecordForm, form_filter
from .models import Records
from django.http import HttpResponse
from openpyxl import Workbook


# Create your views here.

def home(request) :
    
    #Check to see if logging in 
    
    records = Records.objects.all()

    
    if request.method == 'POST' :
        print("Entrei no post")
        username = request.POST['username']
        password = request.POST['password']
        
        #Authenticate
        
        user = authenticate(request, username=username, password=password)
      
        if user is not None :
            login(request,user)
            messages.success(request,"You Have Been Logged In !")
            return redirect('home')
        
        else :
            messages.success(request,"There Was An Error Logging In, Please try again ...")
            return redirect('home')
    
    else :
        
        return render(request, 'home.html', {'records': records})

def logout_user(request) :
   
   logout(request)
   messages.success(request, "You have Been Logged Out ...")
   return redirect('home') 

def register_user(request) :
    
   if request.method == 'POST' :
       
       form = SignUpForm(request.POST)
       if form.is_valid() :
           form.save()
           #Authenticate and login
           
           username = form.cleaned_data['username']
           password = form.cleaned_data['password1']
           user = authenticate(username=username, password=password)
           login(request,user)
           messages.success(request, "You Have Successfully Registered ! Welcome !")
           return redirect('home')
   else :
        form = SignUpForm()
        return render(request, 'register.html', {'form':form})

   return render(request, 'register.html', {'form':form})


def customer_record(request, pk) :
    
    if request.user.is_authenticated:
        #Look up Records
        
        customer_record = Records.objects.get(id=pk)
        return render(request, 'record.html', {'customer_record':customer_record})
 
    else :
       
        messages.success(request, "You Must Be Logged In To View That Page...")
        return redirect('home')

def delete_record(request, pk) :
    
    if request.user.is_authenticated :
        
        delete_it = Records.objects.get(id=pk)
        delete_it.delete()
        messages.success(request, "Record Deleted Successfully")
        return redirect('home')
    
    else :
        
        messages.success(request, "You Must Be Logged In To Do That...")
        return redirect('home')
    
    
def add_record(request) :
    form = AddRecordForm(request.POST or None)
    if request.user.is_authenticated :
        if request.method == "POST" :
            if form.is_valid() :
                add_record = form.save()
                messages.success(request, "Record Added ...")
                return redirect('home')
    
        return render(request, 'add_record.html', {'form':form})
    else : 
        
        messages.success(request, "You Must Be Logged In ...")
        return redirect('home')
    
def update_record(request,pk) :
     if request.user.is_authenticated:
         current_record = Records.objects.get(id=pk)
         form = AddRecordForm(request.POST or None, instance=current_record)
         if form.is_valid() :
             form.save()
             messages.success(request, "Record Has Been Updated !")
             return redirect('home')
         
         return render(request, 'update_record.html', {'form':form})
     else : 
         
        messages.success(request, "You Must Be Logged In ...")
        return redirect('home')
    
def filter_registers(request):
    
    req = {}
    
    if request.method == 'POST':
        
        form = form_filter(request.POST)
        
        if form.is_valid():
            
            field = form.cleaned_data['field'] 
            search = form.cleaned_data['search']
          
            cont = 0
   
            if request.user.is_authenticated :
                
                result = Records.objects.all()
                
                if field =='first_name' :
                    
                    for item in result :
                        if search in item.first_name :
                           cont+=1
                    
                if field =='address' :
                    
                    for item in result :
                        if search in item.address :
                           cont+=1
                
                if field =='last_name' :
                    
                    for item in result :
                        if search in item.last_name :
                            cont+=1
                
                messages.success(request, f"Foram encontrados {cont} registros")
                return render(request, 'filter.html', {'form': form , 'result':result , 'field':field, 'search':search})
            
            else :
                
                messages.success(request, "You Must Be Logged In To Do That...")
                return redirect('home')
    else :
        
        form = form_filter()
    
    
    return render(request, 'home.html', {})

def gerar_excel(request, field, search):
   
    # Filtrar os dados desejados
    resultados = Records.objects.all()
    
    # Criar o arquivo Excel
    workbook = Workbook()
    sheet = workbook.active
    
    # Definir cabeçalhos
    sheet['A1'] = 'First Name'
    sheet['B1'] = 'Last Name'
    sheet['C1'] = 'Email'
    sheet['D1'] = 'Address'
    sheet['E1'] = 'City'
    sheet['F1'] = 'State'
    sheet['G1'] = 'Zipcode'
    sheet['H1'] = 'ID'
    
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15
    
    # Preencher dados
    row_num = 2
    for resultado in resultados:
      if field == 'first_name' and search in resultado.first_name or field == 'last_name' and search in resultado.last_name or field == 'address' and search in resultado.address : 
        sheet.cell(row=row_num, column=1, value=resultado.first_name)
        sheet.cell(row=row_num, column=2, value=resultado.last_name)
        sheet.cell(row=row_num, column=3, value=resultado.email)
        sheet.cell(row=row_num, column=4, value=resultado.address)
        sheet.cell(row=row_num, column=5, value=resultado.city)
        sheet.cell(row=row_num, column=6, value=resultado.state)
        sheet.cell(row=row_num, column=7, value=resultado.zipcode)
        sheet.cell(row=row_num, column=8, value=resultado.pk)
        row_num += 1

    # Configurar o nome do arquivo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio.xlsx"'

    # Salvar o arquivo no response
    workbook.save(response)

    return response
    

